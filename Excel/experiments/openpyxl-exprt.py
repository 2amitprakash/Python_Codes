from openpyxl import Workbook
from openpyxl import load_workbook
# Import necessary style classes
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors

wb = Workbook()
 
# grab the active worksheet
ws = wb.active
 
# Data can be assigned directly to cells
ws['A1'] = 42
 
# Rows can also be appended
ws.append([1, 2, 3])
 
# Save the file
wb.save('sample.xlsx')

######### Append a file
# Start by opening the spreadsheet and selecting the main sheet
workbook = load_workbook(filename="hello.xlsx")
sheet = workbook.active

# Write what you want into a specific cell
sheet["C10"] = "writing ;)"

# Save the spreadsheet
workbook.save(filename="hello.xlsx")

############ Change Style
# Create a few styles
bold_font = Font(bold=True)
big_blue_text = Font(color=colors.BLUE, size=20)
center_aligned_text = Alignment(horizontal="center")
double_border_side = Side(border_style="double")
square_border = Border(top=double_border_side,
                        right=double_border_side,
                        bottom=double_border_side,
                        left=double_border_side)

# Style some cells!
sheet["A2"].font = bold_font
sheet["A3"].font = big_blue_text
sheet["A5"].border = square_border

# Save the spreadsheet
workbook.save(filename="hello.xlsx")